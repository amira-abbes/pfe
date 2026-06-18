import pandas as pd
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape #escape sert à sécuriser le texte avant de l’écrire dans le PDF.
 
 #openpyxl sert à créer le fichier Excel formaté data_quality_tables.xlsx.
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
#reportlab sert à créer le PDF data_quality_report.pdf.
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


#on définit les chemins.
BASE_DIR = Path(__file__).resolve().parents[1]

PROCESSED_DIR = BASE_DIR / "data" / "processed"
REPORTS_DIR = BASE_DIR / "reports"

DATASET_FILE = PROCESSED_DIR / "merged_dataset_inner.csv"
PDF_REPORT_FILE = REPORTS_DIR / "data_quality_report.pdf"
FORMATTED_TABLES_FILE = REPORTS_DIR / "data_quality_tables.xlsx"

#Cette fonction lit un CSV de manière robuste.
def read_csv_safely(path: Path) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1", "ISO-8859-1"]
    separators = [",", ";", "|", "\t"]

    last_error = None

    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    for encoding in encodings:
        for sep in separators:
            try:
                df = pd.read_csv(path, encoding=encoding, sep=sep, engine="python")
                if df.shape[1] > 1:
                    print(
                        f"Lecture réussie : {path.name} | "
                        f"encoding={encoding} | sep='{sep}' | "
                        f"lignes={df.shape[0]} | colonnes={df.shape[1]}"
                    )
                    return df
            except Exception as exc:
                last_error = exc

    raise RuntimeError(
        f"Impossible de lire correctement le fichier {path}. "
        f"Dernière erreur : {last_error}"
    )

"""Cette fonction fait deux choses à la fois :
Elle affiche le message dans le terminal.
Elle ajoute le même message dans la liste lines.
Cette liste sera utilisée ensuite pour créer le rapport texte et le PDF."""
def log_line(lines: list[str], message: str = ""):
    print(message)
    lines.append(message)

#Fonction write_pdf_report
#Cette fonction transforme le rapport texte en PDF.
def write_pdf_report(lines: list[str], output_file: Path):
  
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AuditTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#12355B"),
        spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "AuditSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1F6F8B"),
        spaceBefore=8,
        spaceAfter=5,
    )
    body_style = ParagraphStyle(
        "AuditBody",
        parent=styles["BodyText"],
        fontName="Courier",
        fontSize=8,
        leading=10,
        spaceAfter=2,
    )
#Cette partie crée la structure du PDF au format A4.
    doc = SimpleDocTemplate(
        str(output_file),
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title="Audit qualite Machine Learning Bad Debts",
    )

    story = [Paragraph("Audit qualite du dataset Machine Learning Bad Debts", title_style)]
    for line in lines:
        clean_line = str(line).replace("’", "'")
        stripped = clean_line.strip()
        if not stripped or set(stripped) == {"="}:
            story.append(Spacer(1, 5))
        elif stripped[:2].split(".")[0].isdigit():
            story.append(Paragraph(escape(stripped), section_style))
        else:
            story.append(Paragraph(escape(clean_line), body_style))

    try:
        doc.build(story)
        return output_file
    except PermissionError:#Si le PDF est déjà ouvert dans un lecteur PDF, Windows peut le bloquer.
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_file = output_file.with_name(f"{output_file.stem}_{timestamp}{output_file.suffix}")
        doc.filename = str(fallback_file)
        doc.build(story)
        print(f"Fichier PDF verrouille, copie creee : {fallback_file}") #le script crée une copie avec un nom contenant la date et l’heure
        return fallback_file


def add_dataframe_sheet(workbook: Workbook, sheet_name: str, title: str, description: str, df: pd.DataFrame):
    """
    Ajoute une feuille Excel lisible avec titre, description, tableau, filtres et colonnes ajustees.
    """
    ws = workbook.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="12355B")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    max_cols = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)

    ws["A2"] = description
    ws["A2"].font = Font(italic=True, color="475569")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)

#Si le tableau est vide, le script écrit quand même une feuille Excel avec un message
    table_df = df.copy()
    if table_df.empty:
        table_df = pd.DataFrame({"information": ["Aucune donnee a afficher"]})

    for row in dataframe_to_rows(table_df, index=False, header=True):
        ws.append(row)

    header_row = 3
    thin_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="0F172A")
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 13), 42)

#Elle crée le fichier Excel final.
"""Elle ajoute deux feuilles :
"Valeurs manquantes"
"Controles qualite"""
def write_formatted_tables(
    missing_report: pd.DataFrame,
    quality_checks: pd.DataFrame,
    output_file: Path,
):
    """
    Regroupe les rapports qualite dans un fichier Excel facile a lire.
    """
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # missing_report a les noms des colonnes en index.
    # reset_index() transforme cet index en vraie colonne Excel nommee "column".
    missing_table = missing_report.reset_index().rename(columns={"index": "column"})

    add_dataframe_sheet(
        workbook,
        "Valeurs manquantes",
        "Rapport des valeurs manquantes",
        "Nombre et pourcentage de valeurs absentes pour chaque colonne du dataset fusionne.",
        missing_table,
    )
    add_dataframe_sheet(
        workbook,
        "Controles qualite",
        "Synthese des controles qualite",
        "La colonne statut indique OK ou ATTENTION. La colonne detail explique le controle en langage simple.",
        quality_checks,
    )

    try:
        workbook.save(output_file)
        return output_file
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_file = output_file.with_name(f"{output_file.stem}_{timestamp}{output_file.suffix}")
        workbook.save(fallback_file)
        print(f"Fichier Excel verrouille, copie creee : {fallback_file}")
        return fallback_file


def main():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    lines = []

    log_line(lines, "=" * 90)
    log_line(lines, "Audit qualité du dataset Machine Learning Bad Debts")
    log_line(lines, "=" * 90)

    df = read_csv_safely(DATASET_FILE)

    # check_rows contient les lignes qui seront affichees dans la feuille Excel
    # "Controles qualite".
    #
    # Pour chaque controle, on enregistre 4 informations :
    # - controle : le nom du test effectue.
    # - valeur   : le resultat numerique du test.
    # - statut   : interpretation rapide du resultat.
    #              OK = le controle ne detecte pas de probleme.
    #              ATTENTION = le controle detecte un point a verifier.
    # - detail   : explication simple de pourquoi ce controle est important.
    check_rows = []

    log_line(lines, "\n1. Dimensions du dataset")
    log_line(lines, f"Lignes  : {df.shape[0]}")
    log_line(lines, f"Colonnes: {df.shape[1]}")
    check_rows.append({
        "controle": "Dataset non vide",
        "valeur": df.shape[0],
        # Metrique : nombre de lignes du dataset, df.shape[0].
        # Regle :
        # - OK si nombre de lignes > 0.
        # - ATTENTION si nombre de lignes = 0.
        # Justification : un dataset vide ne peut pas etre audite ni utilise en ML.
        "statut": "OK" if df.shape[0] > 0 else "ATTENTION",
        # detail explique ce que represente la valeur affichee dans Excel.
        "detail": "Nombre de lignes dans le dataset fusionne.",
    })

    log_line(lines, "\n2. Colonnes disponibles")
    for col in df.columns:
        log_line(lines, f" - {col}")

    log_line(lines, "\n3. Types des colonnes")
    dtypes_text = df.dtypes.astype(str)
    for col, dtype in dtypes_text.items():
        log_line(lines, f" - {col}: {dtype}")

    log_line(lines, "\n4. Doublons")
    duplicated_rows = df.duplicated().sum()
    duplicated_msisdn = df["MSISDN"].duplicated().sum() if "MSISDN" in df.columns else "MSISDN absent"

    log_line(lines, f"Lignes totalement dupliquées : {duplicated_rows}")
    log_line(lines, f"Doublons MSISDN              : {duplicated_msisdn}")
    check_rows.append({
        "controle": "Lignes dupliquees",
        "valeur": duplicated_rows,
        # Metrique : nombre de lignes totalement dupliquees, df.duplicated().sum().
        # Regle :
        # - OK si duplicated_rows = 0.
        # - ATTENTION si duplicated_rows > 0.
        # Justification : une ligne repetee peut donner plus de poids a un client
        # ou a une situation dans les calculs.
        "statut": "OK" if duplicated_rows == 0 else "ATTENTION",
        # detail explique le risque : les doublons peuvent fausser les calculs.
        "detail": "Une ligne totalement dupliquee peut fausser les statistiques.",
    })
    check_rows.append({
        "controle": "Doublons MSISDN",
        "valeur": duplicated_msisdn,
        # Metrique : nombre de MSISDN dupliques, df["MSISDN"].duplicated().sum().
        # Regle :
        # - OK si duplicated_msisdn = 0.
        # - ATTENTION si duplicated_msisdn > 0.
        # Justification : MSISDN est la cle client ; un doublon signifie qu'un
        # meme client apparait plusieurs fois dans le fichier fusionne.
        "statut": "OK" if duplicated_msisdn == 0 else "ATTENTION",
        # detail rappelle que MSISDN est la cle d'identification client.
        "detail": "MSISDN doit identifier un client unique dans le fichier fusionne.",
    })

    log_line(lines, "\n5. Valeurs manquantes")
    missing = df.isna().sum()
    missing_pct = (missing / len(df) * 100).round(2)

    # Tableau envoye vers la feuille Excel "Valeurs manquantes".
    # missing_count = nombre de cellules vides par colonne.
    # missing_percent = pourcentage de cellules vides par colonne.
    missing_report = pd.DataFrame({
        "missing_count": missing,
        "missing_percent": missing_pct
    }).sort_values(by="missing_count", ascending=False)

    missing_nonzero = missing_report[missing_report["missing_count"] > 0]

    if missing_nonzero.empty:
        log_line(lines, "Aucune valeur manquante détectée.")
    else:
        log_line(lines, str(missing_nonzero))
    check_rows.append({
        "controle": "Colonnes avec valeurs manquantes",
        "valeur": len(missing_nonzero),
        # Metrique : nombre de colonnes ayant au moins une valeur manquante.
        # Regle :
        # - OK si len(missing_nonzero) = 0.
        # - ATTENTION si len(missing_nonzero) > 0.
        # Justification : les valeurs manquantes doivent etre identifiees avant
        # le notebook pour eviter des erreurs ou des choix de traitement implicites.
        "statut": "OK" if missing_nonzero.empty else "ATTENTION",
        # detail explique pourquoi les valeurs manquantes sont importantes avant le ML.
        "detail": "Les valeurs manquantes doivent etre connues avant le notebook ML.",
    })

    log_line(lines, "\n6. Contrôle des valeurs négatives sur variables financières")
    financial_cols = [
        "AVG_CREDIT_AMOUNT",
        "AVG_CREDIT_FEE",
        "AVG_REIMBURSED_AMOUNT",
        "AVG_FEE_REIMBURSED",
        "AVG_REIMBURSE_RATIO",
        "AVG_DAYS_SINCE_CREDIT",
        "TOTAL_OUTSTANDING_AMOUNT",
        "TOTAL_OUTSTANDING_FEE",
        "NB_SOS",
    ]

    for col in financial_cols:
        if col in df.columns:
            values = pd.to_numeric(df[col], errors="coerce")
            negative_count = (values < 0).sum()
            log_line(lines, f"{col}: valeurs négatives = {negative_count}")
            check_rows.append({
                "controle": f"Valeurs negatives - {col}",
                "valeur": int(negative_count),
                # Metrique : nombre de valeurs negatives dans la colonne financiere.
                # Regle :
                # - OK si negative_count = 0.
                # - ATTENTION si negative_count > 0.
                # Justification : les montants, frais, delais et nombres SOS sont
                # attendus positifs ou nuls dans ce contexte metier.
                "statut": "OK" if negative_count == 0 else "ATTENTION",
                # detail donne la justification metier du controle.
                "detail": "Les variables financieres ne doivent pas contenir de valeurs negatives.",
            })

    log_line(lines, "\n7. Contrôle des ratios")
    if "AVG_REIMBURSE_RATIO" in df.columns:
        ratio = pd.to_numeric(df["AVG_REIMBURSE_RATIO"], errors="coerce")
        below_zero = (ratio < 0).sum()
        above_one = (ratio > 1).sum()

        log_line(lines, f"AVG_REIMBURSE_RATIO < 0 : {below_zero}")
        log_line(lines, f"AVG_REIMBURSE_RATIO > 1 : {above_one}")
        check_rows.append({
            "controle": "AVG_REIMBURSE_RATIO hors intervalle",
            "valeur": int(below_zero + above_one),
            # Metrique : nombre de ratios hors intervalle [0, 1].
            # Regle :
            # - OK si below_zero + above_one = 0.
            # - ATTENTION si below_zero + above_one > 0.
            # Justification : un ratio de remboursement doit etre compris entre
            # 0 et 1, avec 0 = rien rembourse et 1 = totalement rembourse.
            "statut": "OK" if below_zero == 0 and above_one == 0 else "ATTENTION",
            # detail explique la regle metier : 0 = rien rembourse, 1 = totalement rembourse.
            "detail": "Le ratio de remboursement doit rester entre 0 et 1.",
        })

    log_line(lines, "\n8. Contrôle des dates d’activation")
    if "ACCOUNT_ACTIVATED_DATE" in df.columns:
        activation_dates = pd.to_datetime(df["ACCOUNT_ACTIVATED_DATE"], errors="coerce")
        invalid_dates = activation_dates.isna().sum()
        min_date = activation_dates.min()
        max_date = activation_dates.max()

        log_line(lines, f"Dates invalides ou manquantes : {invalid_dates}")
        log_line(lines, f"Date activation minimale      : {min_date}")
        log_line(lines, f"Date activation maximale      : {max_date}")
        check_rows.append({
            "controle": "Dates activation invalides ou manquantes",
            "valeur": int(invalid_dates),
            # Metrique : nombre de dates absentes ou non convertibles en date.
            # Regle :
            # - OK si invalid_dates = 0.
            # - ATTENTION si invalid_dates > 0.
            # Justification : la date d'activation sert a comprendre l'anciennete
            # client ; une date invalide peut perturber les features temporelles.
            "statut": "OK" if invalid_dates == 0 else "ATTENTION",
            # detail donne le contexte temporel observe dans le fichier.
            "detail": f"Periode observee : {min_date} -> {max_date}.",
        })

    # quality_checks devient la feuille Excel "Controles qualite".
    # Colonnes :
    # - controle : nom du test effectue.
    # - valeur : resultat numerique du test.
    # - statut : OK ou ATTENTION.
    # - detail : explication simple du test.
    quality_checks = pd.DataFrame(check_rows)

    log_line(lines, "\n9. Conclusion")
    log_line(lines, "Audit qualité terminé. Les contrôles essentiels ont été générés dans machine_learning/reports.")

    pdf_report_file = write_pdf_report(lines, PDF_REPORT_FILE)
    formatted_tables_file = write_formatted_tables(missing_report, quality_checks, FORMATTED_TABLES_FILE)

    log_line(lines, "\n" + "=" * 90)
    log_line(lines, f"Rapport PDF sauvegardé : {pdf_report_file}")
    log_line(lines, f"Tableaux formatés Excel : {formatted_tables_file}")
    log_line(lines, "=" * 90)


if __name__ == "__main__":
    main()
